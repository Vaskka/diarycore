import json
import os
import re
from openpyxl import load_workbook

from dal import es, mysql
from utils.util import *

maybe_error = []


def excel_rd(path):
    wb = load_workbook(filename=path)
    ws = wb.active

    res = []
    # decide col
    col = {}
    col_i_field = {}
    for i in range(1, 9):
        cell = ws.cell(row=1, column=i).value

        if cell == '日记名':
            col['diaty_title'] = i
            col_i_field[i] = 'diaty_title'
        elif cell == '卷册名':
            col['seq'] = i
            col_i_field[i] = 'seq'
        elif cell == '时间':
            col['diary_date_timestamp'] = i
            col_i_field[i] = 'diary_date_timestamp'
        elif cell == '原文时间or标题':
            col['sub_title'] = i
            col_i_field[i] = 'sub_title'
        elif cell == '日记内容':
            col['content'] = i
            col_i_field[i] = 'content'
        elif cell == '注释':
            col['comment'] = i
            col_i_field[i] = 'comment'
        elif cell == '起始页':
            col['start_end_page'] = i
            col_i_field[i] = 'start_end_page'
        elif cell == '备注':
            col['ext'] = i
            col_i_field[i] = 'ext'
        else:
            raise RuntimeError('非法列:' + cell)

    i = 0
    for rows in ws.rows:
        if i == 0:
            i += 1
            continue

        item = {}
        j = 0
        for cell in rows:
            item[col_i_field[j + 1]] = cell.value
            j += 1
            pass

        res.append(item)
        i += 1

    return res


def detect_author(file_name):
    m = re.search('^.*《(.*)日记.*$', file_name)
    return m.group(1)


def author_exist_or_create(author_name):
    if not mysql.select_by_author_name(author_name=author_name):
        # init author 
        mysql.insert_mysql_author(author_name=author_name, author_avatar_url='', extern_param='{}')
        # relate author type
        mysql.insert_mysql_author_type(author_name=author_name, type_name='未分类')

    return mysql.select_by_author_name(author_name=author_name)['id']


def insert_data(author_id, raw_item):
    # insert db and es
    print('==================', raw_item['diary_date_timestamp'])
    if raw_item['diary_date_timestamp'] is None or raw_item['diary_date_timestamp'] == '':
        print('time not exist,skip,{},{}', author_id, raw_item)
        return
    timestamp = str2timestamp(raw_item['diary_date_timestamp'])
    start_page = get_start_page(raw_item['start_end_page'])
    end_page = get_end_page(raw_item['start_end_page'])
    ext = raw_item['ext']
    ext_param = {
        'ext': ext,
    }
    comment = '' if raw_item['comment'] is None else raw_item['comment']
    # detect error timestamp
    diary_record = mysql.select_by_author_id_title_timestamp(author_id=author_id, title=raw_item['diaty_title'],
                                                             timestamp=timestamp)
    if diary_record is not None:
        maybe_error.append({
            'authorId': author_id,
            'title': raw_item['diaty_title'],
            'sub_title': raw_item['sub_title'],
        })
        return
        pass

    diary_record = mysql.select_by_author_id_title_sub_title(author_id=author_id, title=raw_item['diaty_title'],
                                                             sub_title=raw_item['sub_title'])
    if diary_record is None:
        mysql.insert_mysql_diary(author_id=author_id,
                                 title=raw_item['diaty_title'],
                                 sub_title=raw_item['sub_title'],
                                 timestamp=timestamp,
                                 start_page=start_page,
                                 end_page=end_page,
                                 comment=comment,
                                 ext_param=json.dumps(ext_param))
        diary_record = mysql.select_by_author_id_title_sub_title(author_id=author_id, title=raw_item['diaty_title'],
                                                                 sub_title=raw_item['sub_title'])
    else:
        print('mysql record exist:', raw_item['diaty_title'] + '_' + raw_item['sub_title'])
    diary_id = diary_record['id']

    if not es.exists(diary_id):
        es.insert_es(id=diary_id, diary_id=diary_id, author_id=author_id, timestamp=timestamp,
                     title=raw_item['diaty_title'], sub_title=raw_item['sub_title'], content=raw_item['content'],
                     comment=comment)
    else:
        print('es record exist:', diary_id)
        es.del_es(id=diary_id)
        es.insert_es(id=diary_id, diary_id=diary_id, author_id=author_id, timestamp=timestamp,
                     title=raw_item['diaty_title'], sub_title=raw_item['sub_title'], content=raw_item['content'],
                     comment=comment)
    print('process done,{},{}', author_id, raw_item)
    pass


def process(file_name):
    raw_excel_li = excel_rd(file_name)
    author_name = detect_author(file_name=file_name)

    author_id = author_exist_or_create(author_name=author_name)
    print('process,id={},name={}'.format(author_id, author_name))

    for excel_obj in raw_excel_li:
        insert_data(author_id, excel_obj)

    pass


def main():
    d = './excel'
    for filepath, dirnames, filenames in os.walk(d):
        for filename in filenames:
            path = d + os.sep + filename
            if '.xlsx' in filename and not filename.startswith('~$'):
                print('now process:', path)
                process(path)
            else:
                print('not excel, skip:', path)

    print('all maybe error:', maybe_error)
    pass


def debug():
    # print(len(es.all()))
    # es.del_es(id=7)
    pass


if __name__ == '__main__':
    mysql.setMysqlMock(False)
    es.setEsMock(False)
    if not es.exists_index():
        es.create_index()
    main()
    # debug()
